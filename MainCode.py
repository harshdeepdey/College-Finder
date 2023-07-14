import openpyxl #This library was imported to read the excel sheets

wb=openpyxl.load_workbook("jossa2021.xlsx")#Now 'wb' contains all the information about the excel file 'jossa2021'
'''"jossa2021" consists of 12 excel sheets, where each sheet consists the last year's information about, the opening and the closing rank of the institutes. Here 6 sheets (namely iit1,iit2,iit3....iit6) deals with the rank of adv..Where as other 6 sheets deals with the jee mains rank. '''
#The three strings below(i.e a1,a2 and a3) are defined to pick up a unique excel sheet, according to user's requirement
print("If you want to check for mains type 'main' and for adv type 'adv'")
a1=str(input())#This input helps us to know whether the user is willing to predict the college, based on jee mains rank or based on jee adv rank
print("Select round")
a2=int(input())
a3=a1+str(a2)

sh1=wb[a3]#Now this variable consists all the information present in the excel 
''''''
sheet_cells=[]#This list was defned to store all the information present in excel sheet in form of tuple.
#The for loop below go through each and every row of the sheet, and append the info present in the row in the list 'row_cells' which is later converted into a tuple and is append in the main list 'sheet_cells'
for row in sh1.iter_rows():
    row_cells=[]#This empty list is defined to store the all information present in a unique row
    #The for loop below helps us to achieve the above task
    for cell in row:#This for loop  goes through every cell in a row and append the info present in that cell to an empty list row_cells
        row_cells.append(cell.value)
    sheet_cells.append(tuple(row_cells))#lastly the 'row_cells' is converted into a tuple and was appended into the main lsit "sheet_cells"

print('Enter your category rank ')

rank=int(input())#"Rank" of the user is taken as an input here

'''The next 14 lines of code helps to print all the available colleges for the user'''

d3={}#This dictionary contais all the names of the available colleges corresponding to a number.
z=0#Here z helps to keep track of that corresponding number in the dictionary
ic=[]#This list is defined for keeping the track of, range of rows of all the college(for ex range of IIT bhubaneswar is (2,118)). This is later used to append info in the "d4 dictionary"
for i in range(1,len(sheet_cells)-1):#This for loop goes through every row of the column 2 and on finding a unique name of the institute it appends that name in the dictionary d3
    #The x and y strings compare two consecutive rows , and when x and y is not equal it confirms the fact that the "str x" got an unique name of the institue which is not presnt in d3
    y=str(sh1.cell(i,2).value)
    x=str(sh1.cell(i+1,2).value)
    if(x!=y):
        #On finding the unique name of the institute that name is appended in d3 corresponding to z
        z=z+1
        ic.append(i+1)
        d3[z]=x
        print(z,x,sep=".)")#That name is also printed corresponding to z so that user can choose there desired college just by entering a number
# The option of "All" is also appended in the dictionary d3
z=z+1
print(z,"All",sep=".)")
d3[z]="All"
#The informations appended below in "ic" is later used in appending info in d4
ic.append(len(sheet_cells))
ic.append(2)
ic.append(len(sheet_cells))

print('Enter college type')
print('Select number corresponding to your choice')
collage=int(input())

'''The next 14 lines of code helps to print all the available branches for the user'''
d4={}#This dictionary contais all the names of the available branches in the college the useer choosed, corresponding to a number.
l1=[]#this list keeps track of all the branch names appended in d4
l1.append(str(sh1.cell(ic[collage-1],3).value))
d4[1]=l1[0]
z1=1#Here z1 helps to keep track of that corresponding number in the dictionary
print(z1,l1[0],sep=".)")

#The if else ladder below helps to set the range of the for loop(ie: (a,b)) using the information present in list "ic".
if(collage==z):#If the user chooses the "All" option in colllege the for loop should run from the start of the sheet to the end of the sheet. As the last 2 elements of the list "ic" corresponds to the start of the sheet and the end of the sheet,"a" and "b" are given those values only
    a=ic[len(ic)-2]
    b=ic[len(ic)-1]+1
elif(collage!=z):#If the user chooses any other college it sets the range accordingly
    a=ic[collage-1]
    b=ic[collage]

for i in range(a,b):
    y=str(sh1.cell(i,3).value)
    k=0#This used as a flag variable
    for j in range(len(l1)):#This for loop and the if statemant inside it checks whether the name of the branch is alreadry in the dictionary or not
    #This for loop goes through all the elements of l1 and comapers it with y. If any element in l1 matches with y it assign the value 1 to 'k' (flag variable) and break the loop to continue the same process with another y.
        x=l1[j]
        
        if(x==y):
            k=1
            break
        elif(x!=y):
            k=0
    if(k==0):#As k=0 ensures the uniqueness of the branch name it is appened in d4 corressponding to z1    
        z1=z1+1
        l1.append(y)
        d4[z1]=y
        print(z1,y,sep=".)")
# For user's convenience ,he option of "All" is also appended in the dictionary d4           
z1=z1+1
print(z1,"All",sep=".)")
d4[z1]="All"

print('Enter branch')
print('Select number corresponding to your choice')
branch=int(input())

d5={}#This dictionary contais all the names of the available seat type corresponding to a number.
l2=[]#this list keeps track of all the seat types names appended in d5
l2.append(str(sh1.cell(1,5).value))
z2=0#Here z2 helps to keep track of that corresponding number in the dictionary
k1=0#This used as a flag variable
for i in range(2,len(sheet_cells)):
    y=str(sh1.cell(i,5).value) 
    for j in range(len(l2)):#This for loop and the if statemant inside it checks whether the name of the seat type is alreadry in the dictionary or not
       #This for loop goes through all the elements of l2 and comapers it with y. If any element in l1 matches with y it assign the value 1 to 'k1' (flag variable) and break the loop to continue the same process with another y.
        x=l2[j]     
        if(x==y):
            k1=1
            break
        elif(x!=y):
            k1=0
    if(k1==0):#As k1=0 ensures the uniqueness of the branch name it is appened in d5 corressponding to z2
        z2=z2+1
        l2.append(y)
        d5[z2]=y
        print(z2,y,sep=".)")
    k1=0       
# For user's convenience ,he option of "All" is also appended in the dictionary d5
z2=z2+1
print(z2,"All",sep=".)")
d5[z2]="All"

print('Enter seat type')
print('Select number corresponding to your choice')
seat_type=int(input())#"seat-type" of the user is taken as an input here

print('Enter your gender')
print("If you are 'Gender-Neutral' Type 'G' or if you are 'Female' type 'F'")
gender=str(input())#"Gender" of the user is taken as an input here

w=0
p=0

'''following loops and if statements cover entire permutations and combinations to filter collages and corresponding branches on the basis
 of rank, branch, gender, and type of seat a user is applying for'''

#here we recommend the collage and branch on the basis of rank and gender
#this if statement selects all the collages from given list
if collage==z :
    #this if statement selects all the branches from given list corresponding to the collages
        if branch ==z1:
            #this if selects all the type of seats
            if seat_type ==z2:
                #here we compare the rank intered with the closing rank and also selects only those result whose gender matches with the intered gender
                for i in range(1,len(sheet_cells)-1):
                    #this statement compares the gender, and shortlist the tuples accordingly
                    if gender in sheet_cells[i][5]:
                        #for the shortlisted data on the basis of gender, only the selected set of tuples are checked for rank
                        if sheet_cells[i][7]>rank:
                            w=w+1              #if the above conditions satisfy, the value of w is increased by 1
                            #the set of tuples that qualifies above conditions are printed as output
                            print(w,")",sheet_cells[i][1],sheet_cells[i][2],sheet_cells[i][3],sheet_cells[i][4])
                            p=1


#here we recommend the collage and branch on the basis of rank, gender and ttpe of seat 
#this if statement selects all the collage from given list
if collage ==z:
    #this if statement selects all branches from given list corresponding to all the collages
    if branch ==z1:
        #here we check the rank intered by user with the closing rank and also selects only those result whose gender matches with the intered gender
        for i in range(1,len(sheet_cells)-1):
                if d5[seat_type] ==sheet_cells[i][4]:       #choose the tuple that has same value of seat type as chosen by the user
                    if gender in sheet_cells[i][5]:         #compares the gender, and shortlist the tuples accordingly
                        #for the shortlisted tuples on the basis of seat type and gender, only these set of tuples are checked for rank
                        if sheet_cells[i][7]>rank:
                            w=w+1        #if the above conditions satisfy the value of w is increased by 1
                            #the set of tuples that qualifies above conditions are printed as output
                            print(w,")",sheet_cells[i][1],sheet_cells[i][2],sheet_cells[i][3],sheet_cells[i][4])
                            p=1


#when user wants to see only specific collages and branch by filtering every input, following loop gets executed
for i in range(1,len(sheet_cells)-1):
    #this if statement select only those tuples whose collage name matches with the selected collage by the user 
    if d3[collage] == sheet_cells[i][1]:
        #this if statement selects the branch from given list corresponding to the selected collage
        if branch==z1:
            #this if statement selects all type of seats that are offered in selected collage
            if seat_type==z2:
                #compares the gender, and shortlist the tuple accordingly
                if gender in sheet_cells[i][5]:
                    #for the shortlisted data on the basis of collage, branch, seat type and gender, only these set of tuples is checked for rank
                    if sheet_cells[i][7]>rank:
                        w=w+1       #if the above conditions satisfy the value of w is increased by 1
                        #the set of tuple that qualifies above conditions are printed as output
                        print(w,")",sheet_cells[i][1],sheet_cells[i][2],sheet_cells[i][3],sheet_cells[i][4])
                        p=1


#here we recommend the collage and branch on  the basis of rank, branch and gender
#this if statement selects all collage from given list
if collage==z:
    #here we compare the type of seat, branch, gender and rank using this for loop
    for i in range(1,len(sheet_cells)-1):
        if d4[branch] == sheet_cells[i][2]:     #selects the branch from given list corresponding to selected collage
            if seat_type==z2:                   #this statement selects all type of seat
                if gender in sheet_cells[i][5]:     #compares the gender, and shortlist the data accordingly
                    if sheet_cells[i][7]>rank:      #for the shortlisted data on the basis of branch, seat type and gender, only these set of data is checked for rank
                        w=w+1                       #if the above conditions satisfy the value of w is increased by 1
                        #the set of data that qualifies above conditions are printed as output
                        print(w,")",sheet_cells[i][1],sheet_cells[i][2],sheet_cells[i][3],sheet_cells[i][4])
                        p=1


#here we recommend the collage and branch on  the basis of rank, gender, seat type and branch
#this if statement selects all the collage from given list
if collage==z:
    #for the selected set of collages , following for loop is executed to select the desired output
    for i in range(1,len(sheet_cells)-1):
        #if the branch selected by user matches with the branch in the tuple, those tuples are selected
        if d4[branch] == sheet_cells[i][2]:
            #the selected tuples are again checked for the seat_type and the tuples that fulfill this condition are selected
            if d5[seat_type] == sheet_cells[i][4]:
                #again the selected tuples are checked on the basis of gender, and the tuple are selected accordingly
                if gender in sheet_cells[i][5]:
                    #after all the above conditions, the filtered tuples are checked for the rank
                    if sheet_cells[i][7]>rank:
                        w=w+1         #if the above conditions satisfy the value of w is increased by 1
                        #the set of data that qualifies above conditions are printed as output
                        print(w,")",sheet_cells[i][1],sheet_cells[i][2],sheet_cells[i][3],sheet_cells[i][4])
                        p=1


#here we recommend the collage and branch on the basis of rank, gender, seat type and collage
for i in range(1,len(sheet_cells)-1):
    #here we choose only those tuples whose collages matches with the collage selected by the user
    if d3[collage] in sheet_cells[i][1]:
        if branch==z1:      #all the branches that are offered in selected collage are selected by this statement
            #now for the selected tuples the type of seat intered is compared against each tuple
            if d5[seat_type] == sheet_cells[i][4]:
                # the tuples are again eliminated based on the gender intered by the user
                if gender in sheet_cells[i][5]:
                    #finally the shortlisted tuples are checked for rank 
                    if sheet_cells[i][7]>rank:
                        w=w+1        #if the above conditions satisfy the value of w is increased by 1
                        #the set of data that qualifies above conditions are printed as output
                        print(w,")",sheet_cells[i][1],sheet_cells[i][2],sheet_cells[i][3],sheet_cells[i][4])
                        p=1

#in this loop we recommend the collage and branch on  the basis of rank for the selected collage, branch and gender
for i in range(1,len(sheet_cells)-1):
    #here we choose only those tuples whose collages matches with the collage selected by the user
    if d3[collage] in sheet_cells[i][1]:
        #if the branch selected by user matches with the branche in the tuple, those tuples are seleted
        if d4[branch] in sheet_cells[i][2]:
            if seat_type==z2:       #this statement selects all type of seat
                if gender in sheet_cells[i][5]:     # the tuples are again eliminated based on the gender intered by the user
                    if sheet_cells[i][7]>rank:      #finally the shortlisted tuples are checked for rank 
                        w=w+1            #if the above conditions satisfy the value of w is increased by 1
                        #the set of data that qualifies above conditions are printed as output
                        print(w,")",sheet_cells[i][1],sheet_cells[i][2],sheet_cells[i][3],sheet_cells[i][4])
                        p=1


#in this loop we recommend the collage and branch on  the basis of rank for the selected collage, branch, seat type and gender
for i in range(1,len(sheet_cells)-1):
    #here we choose only those tuples whose collages matches with the collage selected by the user
    if d3[collage] in sheet_cells[i][1]:
        #if the branch selected by user matches with the branch in the tuple, those tuples are selected
        if d4[branch] in sheet_cells[i][2]:             # the tuples are eliminated based on the branch selected by the user
            if d5[seat_type] in sheet_cells[i][4]:      # this statement check the seat type and selects tuples accordingly
                if gender in sheet_cells[i][5]:         # the tuples are again eliminated based on the gender intered by the user
                    if sheet_cells[i][7]>rank:          # finally the shortlisted tuples are checked for rank
                        w=w+1            #if the above conditions satisfy, the value of w is increased by 1
                        #the set of data that qualifies above conditions are printed as output
                        print(w,")",sheet_cells[i][1],sheet_cells[i][2],sheet_cells[i][3],sheet_cells[i][4])
                        p=1

#if none of the above condition is fulfilled, that means there is no collage corresponding to your data intered
if(p==0):
    print("Sorry, no college found based on your inputs")
